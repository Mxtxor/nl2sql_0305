"""xlwings 없이 openpyxl로 파싱 로직 검증 (로컬 단위 테스트용)"""
import openpyxl, re

def glossary_preprocessing(texts):
    cleaned = []
    for text in texts:
        if not text:
            continue
        text = str(text).strip()
        text = re.sub(r'\s+', ' ', text)
        if not text:
            continue
        cleaned.append(text)
    return cleaned

def load_glossary_from_xlsx_local(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    raw_data, metadatas = [], []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        term      = str(d.get('standard_term (논리명)')             or '').strip()
        phys      = str(d.get('standard_term_nm (물리명)')          or '').strip()
        desp      = str(d.get('standard_term_desp (표준용어 설명)') or '').strip()
        cd_nm     = str(d.get('cmn_cd_nm (공통코드명)')             or '').strip()
        cd        = str(d.get('cmn_cd (공통코드)')                  or '').strip()
        cd_grp_id = str(d.get('cd_grp_id (코드그룹아이디)')         or '').strip()
        cd_grp_nm = str(d.get('cd_grp_nm (코드그룹명)')             or '').strip()
        col_id    = str(d.get('칼럼ID')                             or '').strip()
        if not term:
            continue
        text = f"{term} ({phys}): {desp}"
        if cd_nm and cd:
            text += f" 코드값: {cd_nm}={cd}"
        raw_data.append(text)
        metadatas.append({
            "term_name": term, "physical_name": phys, "description": desp,
            "cd_grp_id": cd_grp_id, "cd_grp_nm": cd_grp_nm,
            "cmn_cd_nm": cd_nm, "cmn_cd": cd, "column_id": col_id,
        })
    return raw_data, metadatas

if __name__ == "__main__":
    raw, meta = load_glossary_from_xlsx_local("glossary.xlsx")
    cleaned = glossary_preprocessing(raw)
    print(f"\n[파싱 결과] 총 {len(cleaned)}개 항목\n")
    for i, (t, m) in enumerate(zip(cleaned, meta)):
        print(f"--- 항목 {i+1} ---")
        print(f"[raw_data]  {t}")
        print(f"[metadata]  {m}")
        print()
