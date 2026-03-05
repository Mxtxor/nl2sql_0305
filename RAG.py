import os
import json
import boto3
from rag_utils import SchemaIngestConfig, SchemaIngester
from glossary_utils import glossary_preprocessing


def load_glossary_from_xlsx(path: str):
    """
    xlsx 파일을 openpyxl로 파싱하여 (raw_data, metadatas) 형태로 반환한다.

    - raw_data  : OpenSearch에 임베딩될 자연어 문장 리스트 (List[str])
    - metadatas : 각 문장에 대응하는 구조화 메타 딕셔너리 리스트 (List[dict])

    openpyxl 사용 이유:
      Linux 서버(Docker) 환경에서는 Excel 앱이 없어 xlwings 사용 불가.
      openpyxl은 순수 Python 라이브러리로 Linux에서 정상 동작하며,
      서버 환경에서는 사내 보안 프로그램이 적용되지 않으므로 사용 가능.
    """
    import openpyxl

    raw_data = []
    metadatas = []

    wb = openpyxl.load_workbook(os.path.abspath(path))
    ws = wb.active

    # 헤더 행(1행) 읽기 — 1~2행 병합 구조이므로 1행 기준
    headers = [cell.value for cell in ws[1]]

    # 데이터는 3행부터 시작 (1~2행이 병합 헤더)
    data_rows = list(ws.iter_rows(min_row=3, values_only=True))

    if not data_rows:
        return raw_data, metadatas


    for row in data_rows:
        d = dict(zip(headers, row))

        term      = str(d.get('standard_term (논리명)')             or '').strip()
        phys      = str(d.get('standard_term_nm (물리명)')          or '').strip()
        desp      = str(d.get('standard_term_desp (표준용어 설명)') or '').strip()
        cmn_cd_nm = str(d.get('cmn_cd_nm (공통코드명)')             or '').strip()
        cmn_cd    = str(d.get('cmn_cd (공통코드)')                  or '').strip()
        cd_grp_id = str(d.get('cd_grp_id (코드그룹아이디)')         or '').strip()
        cd_grp_nm = str(d.get('cd_grp_nm (코드그룹명)')             or '').strip()
        col_id    = str(d.get('칼럼ID')                             or '').strip()

        # 논리명이 없는 빈 행 스킵
        if not term:
            continue

        # raw_data: 임베딩 대상 자연어 텍스트 구성
        text = f"{term} ({phys}): {desp}"
        if cmn_cd_nm and cmn_cd:
            text += f" 코드값: {cmn_cd_nm}={cmn_cd}"
        raw_data.append(text)

        # metadatas: 검색 결과 활용을 위한 구조화 데이터
        metadatas.append({
            "term_name":     term,
            "physical_name": phys,
            "description":   desp,
            "cd_grp_id":     cd_grp_id,
            "cd_grp_nm":     cd_grp_nm,
            "cmn_cd_nm":     cmn_cd_nm,
            "cmn_cd":        cmn_cd,
            "column_id":     col_id,
        })

    return raw_data, metadatas


def insert_glossary(xlsx_path: str):
    """
    xlsx 파일을 파싱하여 Glossary 데이터를 OpenSearch에 적재한다.
    """
    config = SchemaIngestConfig()
    client = SchemaIngester(config)

    raw_data, metadatas = load_glossary_from_xlsx(xlsx_path)
    print(f"[Glossary] 파싱 완료: {len(raw_data)}개 항목")

    client.ingest_glossary(raw_data, metadatas)


def create_t2s_retriever():
    config = SchemaIngestConfig()
    client = SchemaIngester(config)
    client.ingest_topic(["hello"], metadatas=[
        {"content": "hello"},
        {"table_name": "salary"}
    ])

#def insert_topic():

#def insert_table():

if __name__ == "__main__":
    insert_glossary("glossary.xlsx")